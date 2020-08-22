from django.http import HttpRequest, HttpResponse
from django.shortcuts import render, redirect
from openpyxl import Workbook

from Models.Empleado.forms import FormularioEmpleado
from Models.Empleado.models import Empleado


class FormularioEmpleadoView(HttpRequest):
    def index(request):
        empleado = FormularioEmpleado()
        return render(request, "EmpleadoIndex.html", {"form":empleado})

    def procesar_formulario(request):
        empleado = FormularioEmpleado(request.POST)
        if empleado.is_valid():
            empleado.save()
            empleado = FormularioEmpleado()

            return render(request, "EmpleadoIndex.html", {"form":empleado, "mensaje": 'ok'})

    def listar_empleados(request):
        empleado = Empleado.objects.all()
        return render(request, "ListaEmpleado.html", {"empleados":empleado})

    def eliminarEmpleado(request,id):
        Empleado.objects.filter(CodigoEmpleado=id).delete()
        return redirect(to="listarEmpleado")

    def editarEmpleado(request, id):
        actua= Empleado.objects.get(CodigoEmpleado=id)
        data = {
            'form': FormularioEmpleado(instance=actua)
        }
        if request.method == 'POST':
            formulario = FormularioEmpleado(data=request.POST, instance=actua)
            if formulario.is_valid():
                formulario.save()
                data['mensaje']= "Se Actualizo el Registro."
                data['form']=formulario

        return render(request, 'modificarEmpleado.html',data)

    def ReporteExcelEmpleado(self,*args, **kwargs):
        empleados = Empleado.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = "Reporte de Empleados"

        ws.merge_cells('B1:G1')

        ws['B3'] = 'Codigo de Empleado'
        ws['C3'] = 'Nombre '
        ws['D3'] = 'Apellido '
        ws['E3'] = 'Telefono'
        ws['F3'] = 'DPI'
        ws['G3'] = 'Direccion '

        cont = 6

        for empleados in empleados:
            ws.cell(row = cont, column = 2).value = empleados.CodigoEmpleado
            ws.cell(row = cont, column = 3).value = empleados.Nombre
            ws.cell(row=cont, column=4).value = empleados.Apellido
            ws.cell(row=cont, column=5).value = empleados.Telefono
            ws.cell(row=cont, column=6).value = empleados.DPI
            ws.cell(row=cont, column=7).value = empleados.Direccion
            cont+=1

        nombre_archivo = "ReporteProyectosExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response



